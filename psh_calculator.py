#!/usr/bin/env python3
"""
PSH Rent Calculator - Excel Edition
A user-friendly desktop application for calculating PSH rent assistance
Works entirely with Excel spreadsheets - no internet required!

Built for low-tech users with simple, guided interface.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
from decimal import Decimal, ROUND_DOWN
import json


class PSHCalculatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PSH Rent Calculator - Excel Edition")
        self.root.geometry("800x700")
        
        # Colors for low-tech friendly interface
        self.colors = {
            'primary': '#2E7D32',      # Green - safe/go
            'secondary': '#1976D2',    # Blue - info
            'warning': '#F57C00',      # Orange - attention needed
            'danger': '#D32F2F',       # Red - stop/error
            'success': '#388E3C',      # Dark green - success
            'light_gray': '#F5F5F5',   # Light background
            'white': '#FFFFFF'
        }
        
        # FMR data (can be updated by user)
        self.fmr_data = {
            0: 2485,  # Studio
            1: 2977,  # 1-BR
            2: 3604,  # 2-BR  
            3: 4604,  # 3-BR
            4: 4772,  # 4-BR
            5: 5000   # 5-BR (estimate)
        }
        
        # Load saved settings
        self.load_settings()
        
        # Initialize variables
        self.current_calculation = {}
        
        # Build UI
        self.setup_styles()
        self.build_ui()
        
    def setup_styles(self):
        """Configure ttk styles for better appearance"""
        style = ttk.Style()
        
        # Configure button styles
        style.configure('Primary.TButton',
                       background=self.colors['primary'],
                       foreground='white',
                       font=('Arial', 10, 'bold'))
        
        style.configure('Warning.TButton',
                       background=self.colors['warning'],
                       foreground='white',
                       font=('Arial', 9, 'bold'))
        
        style.configure('Large.TLabel',
                       font=('Arial', 12, 'bold'))
        
        style.configure('Header.TLabel',
                       font=('Arial', 14, 'bold'),
                       foreground=self.colors['primary'])
        
        style.configure('Success.TLabel',
                       foreground=self.colors['success'],
                       font=('Arial', 10, 'bold'))
        
        style.configure('Warning.TLabel',
                       foreground=self.colors['warning'],
                       font=('Arial', 10, 'bold'))
        
        style.configure('Error.TLabel',
                       foreground=self.colors['danger'],
                       font=('Arial', 10, 'bold'))
    
    def build_ui(self):
        """Build the main user interface"""
        # Main container
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, 
                               text="PSH Rent Calculator", 
                               style='Header.TLabel')
        title_label.grid(row=0, column=0, pady=(0, 20))
        
        subtitle_label = ttk.Label(main_frame, 
                                  text="Excel Edition - Works Offline",
                                  font=('Arial', 10, 'italic'))
        subtitle_label.grid(row=1, column=0, pady=(0, 30))
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 20))
        main_frame.rowconfigure(2, weight=1)
        
        # Create tabs
        self.create_calculator_tab()
        self.create_settings_tab()
        self.create_help_tab()
        
        # Status bar
        self.status_var = tk.StringVar()
        self.status_var.set("Ready - Select a tab to begin")
        status_label = ttk.Label(main_frame, textvariable=self.status_var,
                                relief=tk.SUNKEN, anchor=tk.W, padding="5")
        status_label.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
    
    def create_calculator_tab(self):
        """Create the main calculator interface"""
        calc_frame = ttk.Frame(self.notebook, padding="20")
        self.notebook.add(calc_frame, text="📊 Calculator")
        
        # Configure grid
        calc_frame.columnconfigure(1, weight=1)
        
        # Household Information Section
        household_frame = ttk.LabelFrame(calc_frame, text="📋 Household Information", padding="10")
        household_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        household_frame.columnconfigure(1, weight=1)
        
        # Head of Household
        ttk.Label(household_frame, text="Head of Household Name:", font=('Arial', 10, 'bold')).grid(
            row=0, column=0, sticky=tk.W, pady=5)
        self.hoh_var = tk.StringVar()
        hoh_entry = ttk.Entry(household_frame, textvariable=self.hoh_var, width=40)
        hoh_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 0))
        
        # Voucher Size
        ttk.Label(household_frame, text="Voucher Bedroom Size:", font=('Arial', 10, 'bold')).grid(
            row=1, column=0, sticky=tk.W, pady=5)
        self.voucher_var = tk.StringVar(value="0")
        voucher_combo = ttk.Combobox(household_frame, textvariable=self.voucher_var,
                                    values=['0', '1', '2', '3', '4', '5'], 
                                    state='readonly', width=10)
        voucher_combo.grid(row=1, column=1, sticky=tk.W, pady=5, padx=(10, 0))
        voucher_combo.bind('<<ComboboxSelected>>', self.update_fmr_display)
        
        # Bedrooms Leased
        ttk.Label(household_frame, text="Bedrooms in Unit:", font=('Arial', 10, 'bold')).grid(
            row=2, column=0, sticky=tk.W, pady=5)
        self.bedrooms_var = tk.StringVar(value="0")
        bedrooms_combo = ttk.Combobox(household_frame, textvariable=self.bedrooms_var,
                                     values=['0', '1', '2', '3', '4', '5'], 
                                     state='readonly', width=10)
        bedrooms_combo.grid(row=2, column=1, sticky=tk.W, pady=5, padx=(10, 0))
        bedrooms_combo.bind('<<ComboboxSelected>>', self.update_fmr_display)
        
        # FMR Display
        self.fmr_info_var = tk.StringVar()
        self.fmr_info_label = ttk.Label(household_frame, textvariable=self.fmr_info_var,
                                       style='Success.TLabel')
        self.fmr_info_label.grid(row=3, column=0, columnspan=2, sticky=tk.W, pady=(10, 5))
        
        # Financial Information Section
        finance_frame = ttk.LabelFrame(calc_frame, text="💰 Financial Information", padding="10")
        finance_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        finance_frame.columnconfigure(1, weight=1)
        
        # Rent to Owner
        ttk.Label(finance_frame, text="Rent to Owner ($):", font=('Arial', 10, 'bold')).grid(
            row=0, column=0, sticky=tk.W, pady=5)
        self.rent_var = tk.StringVar(value="0")
        rent_entry = ttk.Entry(finance_frame, textvariable=self.rent_var, width=15)
        rent_entry.grid(row=0, column=1, sticky=tk.W, pady=5, padx=(10, 0))
        rent_entry.bind('<KeyRelease>', self.update_calculations)
        
        # Utility Allowance
        ttk.Label(finance_frame, text="Utility Allowance ($):", font=('Arial', 10, 'bold')).grid(
            row=1, column=0, sticky=tk.W, pady=5)
        self.utility_var = tk.StringVar(value="0")
        utility_entry = ttk.Entry(finance_frame, textvariable=self.utility_var, width=15)
        utility_entry.grid(row=1, column=1, sticky=tk.W, pady=5, padx=(10, 0))
        utility_entry.bind('<KeyRelease>', self.update_calculations)
        
        # Total Tenant Payment (TTP)
        ttk.Label(finance_frame, text="Total Tenant Payment ($):", font=('Arial', 10, 'bold')).grid(
            row=2, column=0, sticky=tk.W, pady=5)
        self.ttp_var = tk.StringVar(value="50")
        ttp_entry = ttk.Entry(finance_frame, textvariable=self.ttp_var, width=15)
        ttp_entry.grid(row=2, column=1, sticky=tk.W, pady=5, padx=(10, 0))
        ttp_entry.bind('<KeyRelease>', self.update_calculations)
        
        # Family Composition Section
        family_frame = ttk.LabelFrame(calc_frame, text="👥 Family Composition", padding="10")
        family_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        family_frame.columnconfigure(1, weight=1)
        
        # Eligible Members
        ttk.Label(family_frame, text="Eligible Members:", font=('Arial', 10, 'bold')).grid(
            row=0, column=0, sticky=tk.W, pady=5)
        self.eligible_var = tk.StringVar(value="1")
        eligible_spin = ttk.Spinbox(family_frame, from_=1, to=20, textvariable=self.eligible_var, width=10)
        eligible_spin.grid(row=0, column=1, sticky=tk.W, pady=5, padx=(10, 0))
        eligible_spin.bind('<KeyRelease>', self.update_calculations)
        
        # Ineligible Members
        ttk.Label(family_frame, text="Ineligible Members:", font=('Arial', 10, 'bold')).grid(
            row=1, column=0, sticky=tk.W, pady=5)
        self.ineligible_var = tk.StringVar(value="0")
        ineligible_spin = ttk.Spinbox(family_frame, from_=0, to=20, textvariable=self.ineligible_var, width=10)
        ineligible_spin.grid(row=1, column=1, sticky=tk.W, pady=5, padx=(10, 0))
        ineligible_spin.bind('<KeyRelease>', self.update_calculations)
        
        # Mixed Family Warning
        self.mixed_family_var = tk.StringVar()
        self.mixed_family_label = ttk.Label(family_frame, textvariable=self.mixed_family_var,
                                           style='Warning.TLabel')
        self.mixed_family_label.grid(row=2, column=0, columnspan=2, sticky=tk.W, pady=(10, 5))
        
        # Results Section
        results_frame = ttk.LabelFrame(calc_frame, text="📈 Calculation Results", padding="10")
        results_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        results_frame.columnconfigure(1, weight=1)
        
        # Key Results Display
        self.gross_rent_var = tk.StringVar(value="$0")
        self.hap_to_owner_var = tk.StringVar(value="$0")
        self.tenant_rent_var = tk.StringVar(value="$0")
        self.utility_reimb_var = tk.StringVar(value="$0")
        
        # Results grid
        ttk.Label(results_frame, text="Gross Rent:", font=('Arial', 10, 'bold')).grid(
            row=0, column=0, sticky=tk.W, pady=2)
        ttk.Label(results_frame, textvariable=self.gross_rent_var, font=('Arial', 10)).grid(
            row=0, column=1, sticky=tk.W, pady=2, padx=(10, 0))
        
        ttk.Label(results_frame, text="HAP to Owner:", font=('Arial', 10, 'bold')).grid(
            row=1, column=0, sticky=tk.W, pady=2)
        ttk.Label(results_frame, textvariable=self.hap_to_owner_var, 
                 font=('Arial', 12, 'bold'), foreground=self.colors['success']).grid(
            row=1, column=1, sticky=tk.W, pady=2, padx=(10, 0))
        
        ttk.Label(results_frame, text="Tenant Rent:", font=('Arial', 10, 'bold')).grid(
            row=2, column=0, sticky=tk.W, pady=2)
        ttk.Label(results_frame, textvariable=self.tenant_rent_var, 
                 font=('Arial', 12, 'bold'), foreground=self.colors['primary']).grid(
            row=2, column=1, sticky=tk.W, pady=2, padx=(10, 0))
        
        ttk.Label(results_frame, text="Utility Reimbursement:", font=('Arial', 10, 'bold')).grid(
            row=3, column=0, sticky=tk.W, pady=2)
        ttk.Label(results_frame, textvariable=self.utility_reimb_var, font=('Arial', 10)).grid(
            row=3, column=1, sticky=tk.W, pady=2, padx=(10, 0))
        
        # Warnings/Alerts
        self.alert_var = tk.StringVar()
        self.alert_label = ttk.Label(results_frame, textvariable=self.alert_var,
                                    style='Error.TLabel', wraplength=500)
        self.alert_label.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 5))
        
        # Action Buttons
        button_frame = ttk.Frame(calc_frame)
        button_frame.grid(row=4, column=0, columnspan=2, pady=20)
        
        # Calculate Button
        calc_button = ttk.Button(button_frame, text="📊 Calculate Rent",
                                style='Primary.TButton', command=self.perform_calculation)
        calc_button.pack(side=tk.LEFT, padx=5)
        
        # Export to Excel Button
        export_button = ttk.Button(button_frame, text="📄 Export to Excel",
                                  command=self.export_to_excel)
        export_button.pack(side=tk.LEFT, padx=5)
        
        # Clear Form Button
        clear_button = ttk.Button(button_frame, text="🗑️ Clear Form",
                                 command=self.clear_form)
        clear_button.pack(side=tk.LEFT, padx=5)
        
        # Initial calculation
        self.update_fmr_display()
        self.update_calculations()
    
    def create_settings_tab(self):
        """Create settings tab for FMR management"""
        settings_frame = ttk.Frame(self.notebook, padding="20")
        self.notebook.add(settings_frame, text="⚙️ Settings")
        
        # FMR Settings
        fmr_frame = ttk.LabelFrame(settings_frame, text="🏠 Fair Market Rent (FMR) Rates", padding="10")
        fmr_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # Instructions
        instructions = ttk.Label(fmr_frame, 
                                text="Update FMR rates for your area. These are used to determine payment standards.",
                                wraplength=600)
        instructions.pack(anchor=tk.W, pady=(0, 15))
        
        # FMR entry widgets
        self.fmr_vars = {}
        fmr_grid_frame = ttk.Frame(fmr_frame)
        fmr_grid_frame.pack(anchor=tk.W)
        
        bedroom_labels = ["Studio (0-BR)", "1-Bedroom", "2-Bedroom", "3-Bedroom", "4-Bedroom", "5-Bedroom"]
        
        for i, label in enumerate(bedroom_labels):
            ttk.Label(fmr_grid_frame, text=f"{label}:", font=('Arial', 10, 'bold')).grid(
                row=i, column=0, sticky=tk.W, pady=5, padx=5)
            
            self.fmr_vars[i] = tk.StringVar(value=str(self.fmr_data[i]))
            fmr_entry = ttk.Entry(fmr_grid_frame, textvariable=self.fmr_vars[i], width=15)
            fmr_entry.grid(row=i, column=1, pady=5, padx=(10, 5))
            
            ttk.Label(fmr_grid_frame, text="$", font=('Arial', 10)).grid(
                row=i, column=2, sticky=tk.W, pady=5)
        
        # FMR Buttons
        fmr_button_frame = ttk.Frame(fmr_frame)
        fmr_button_frame.pack(anchor=tk.W, pady=15)
        
        ttk.Button(fmr_button_frame, text="💾 Save FMR Rates",
                  style='Primary.TButton', command=self.save_fmr_rates).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(fmr_button_frame, text="🔄 Reset to Defaults",
                  command=self.reset_fmr_rates).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(fmr_button_frame, text="📁 Load from Excel",
                  command=self.load_fmr_from_excel).pack(side=tk.LEFT, padx=5)
        
        # Data Management
        data_frame = ttk.LabelFrame(settings_frame, text="💾 Data Management", padding="10")
        data_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(data_frame, text="Manage your calculation data and settings.",
                 wraplength=600).pack(anchor=tk.W, pady=(0, 10))
        
        data_button_frame = ttk.Frame(data_frame)
        data_button_frame.pack(anchor=tk.W)
        
        ttk.Button(data_button_frame, text="📤 Export Settings",
                  command=self.export_settings).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(data_button_frame, text="📥 Import Settings",
                  command=self.import_settings).pack(side=tk.LEFT, padx=5)
    
    def create_help_tab(self):
        """Create help and about tab"""
        help_frame = ttk.Frame(self.notebook, padding="20")
        self.notebook.add(help_frame, text="❓ Help")
        
        # Help content
        help_text = """
PSH RENT CALCULATOR - USER GUIDE

GETTING STARTED:
1. Enter household information (name, voucher size, bedrooms in unit)
2. Enter financial details (rent, utility allowance, tenant payment)  
3. Enter family composition (eligible/ineligible members)
4. Click 'Calculate Rent' to see results
5. Click 'Export to Excel' to save a formatted report

KEY FEATURES:
• Works completely offline - no internet required
• Automatic validation and error checking
• Handles mixed families with proration
• Generates professional Excel reports
• Warns about FMR violations requiring supervisor approval

UNDERSTANDING THE RESULTS:
• HAP to Owner: Amount housing authority pays landlord
• Tenant Rent: Amount tenant pays landlord directly
• Utility Reimbursement: Amount tenant receives for utilities
• Gross Rent: Total rent (rent to owner + utility allowance)

MIXED FAMILIES:
When some family members don't have eligible immigration status,
the assistance is prorated based on eligible vs. total members.

FMR (FAIR MARKET RENT):
These are HUD-published rent limits for your area. If gross rent
exceeds FMR, supervisor approval is typically required.

TIPS FOR LOW-TECH USERS:
• All fields show helpful hints and validation
• Results update automatically as you type
• Red warnings indicate issues that need attention
• Green text shows successful calculations
• Use Tab key to move between fields quickly

EXCEL EXPORT:
Creates a professional report with:
• Complete calculation breakdown
• Formatted for easy printing
• Includes all input data and results
• Shows warnings and approval requirements
        """
        
        # Create scrollable text widget
        text_frame = ttk.Frame(help_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(text_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        help_text_widget = tk.Text(text_frame, wrap=tk.WORD, yscrollcommand=scrollbar.set,
                                  font=('Arial', 10), padx=15, pady=15)
        help_text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar.config(command=help_text_widget.yview)
        
        help_text_widget.insert(tk.END, help_text)
        help_text_widget.config(state=tk.DISABLED)
        
        # About section
        about_frame = ttk.LabelFrame(help_frame, text="📋 About", padding="10")
        about_frame.pack(fill=tk.X, pady=15)
        
        about_text = """
PSH Rent Calculator - Excel Edition v1.0
Built for Housing Authorities and PSH Program Staff

• No internet connection required
• Works with standard Excel files  
• Designed for non-technical users
• Follows HUD PSH calculation guidelines
• Professional report generation
        """
        
        ttk.Label(about_frame, text=about_text, justify=tk.LEFT).pack(anchor=tk.W)
    
    def update_fmr_display(self, event=None):
        """Update FMR information display"""
        try:
            voucher_size = int(self.voucher_var.get())
            bedrooms = int(self.bedrooms_var.get())
            
            # Use minimum of voucher size and actual bedrooms for FMR lookup
            fmr_bedrooms = min(voucher_size, bedrooms)
            fmr_amount = self.fmr_data.get(fmr_bedrooms, 0)
            
            if fmr_amount > 0:
                self.fmr_info_var.set(f"FMR for {fmr_bedrooms}-bedroom: ${fmr_amount:,}")
            else:
                self.fmr_info_var.set("FMR information not available")
                
        except (ValueError, KeyError):
            self.fmr_info_var.set("")
        
        # Trigger calculation update
        self.update_calculations()
    
    def update_calculations(self, event=None):
        """Update calculations as user types"""
        try:
            # Get values
            rent = float(self.rent_var.get() or 0)
            utility = float(self.utility_var.get() or 0)
            ttp = max(50, float(self.ttp_var.get() or 50))  # Minimum $50
            eligible = int(self.eligible_var.get() or 1)
            ineligible = int(self.ineligible_var.get() or 0)
            
            # Basic calculations
            gross_rent = rent + utility
            total_hap = gross_rent - ttp
            hap_to_owner = min(rent, max(0, total_hap))
            tenant_rent = max(0, rent - total_hap)
            utility_reimb = max(0, utility - ttp)
            
            # Mixed family calculations
            total_members = eligible + ineligible
            is_mixed = ineligible > 0
            
            if is_mixed:
                prorate_pct = eligible / total_members
                prorated_hap = int(prorate_pct * hap_to_owner)
                mixed_family_rent = rent - prorated_hap
                self.mixed_family_var.set(
                    f"⚠️ MIXED FAMILY: Prorated assistance = {prorate_pct:.1%} "
                    f"(HAP: ${prorated_hap:,}, Tenant: ${mixed_family_rent:,})")
            else:
                self.mixed_family_var.set("✓ All members eligible - no proration needed")
            
            # Update display
            self.gross_rent_var.set(f"${gross_rent:,.0f}")
            self.hap_to_owner_var.set(f"${hap_to_owner:,.0f}")
            self.tenant_rent_var.set(f"${tenant_rent:,.0f}")
            self.utility_reimb_var.set(f"${utility_reimb:,.0f}")
            
            # Check for FMR violation
            voucher_size = int(self.voucher_var.get())
            bedrooms = int(self.bedrooms_var.get())
            fmr_bedrooms = min(voucher_size, bedrooms)
            fmr_amount = self.fmr_data.get(fmr_bedrooms, 0)
            
            if fmr_amount > 0 and gross_rent > fmr_amount:
                over_fmr = gross_rent - fmr_amount
                self.alert_var.set(
                    f"⚠️ WARNING: Gross rent exceeds FMR by ${over_fmr:,.0f}. "
                    f"Supervisor approval required!")
            else:
                self.alert_var.set("")
            
            # Store current calculation
            self.current_calculation = {
                'hoh_name': self.hoh_var.get(),
                'voucher_size': voucher_size,
                'bedrooms': bedrooms,
                'rent_to_owner': rent,
                'utility_allowance': utility,
                'ttp': ttp,
                'eligible_members': eligible,
                'ineligible_members': ineligible,
                'total_members': total_members,
                'is_mixed_family': is_mixed,
                'gross_rent': gross_rent,
                'fmr': fmr_amount,
                'over_fmr': max(0, gross_rent - fmr_amount) if fmr_amount > 0 else 0,
                'total_hap': total_hap,
                'hap_to_owner': hap_to_owner,
                'tenant_rent': tenant_rent,
                'utility_reimbursement': utility_reimb,
                'calculation_date': datetime.now().strftime('%m/%d/%Y'),
                'fmr_bedrooms': fmr_bedrooms
            }
            
            if is_mixed:
                self.current_calculation.update({
                    'prorate_percentage': prorate_pct,
                    'prorated_hap': prorated_hap,
                    'mixed_family_rent': mixed_family_rent
                })
            
            self.status_var.set("Calculations updated automatically")
            
        except (ValueError, ZeroDivisionError):
            # Clear displays on invalid input
            self.gross_rent_var.set("$0")
            self.hap_to_owner_var.set("$0")
            self.tenant_rent_var.set("$0")
            self.utility_reimb_var.set("$0")
            self.alert_var.set("")
            self.mixed_family_var.set("")
            self.status_var.set("Enter valid numbers to see calculations")
    
    def perform_calculation(self):
        """Perform final calculation with validation"""
        # Validate required fields
        if not self.hoh_var.get().strip():
            messagebox.showerror("Validation Error", 
                               "Please enter the Head of Household name")
            return
        
        try:
            rent = float(self.rent_var.get() or 0)
            if rent <= 0:
                messagebox.showerror("Validation Error", 
                                   "Rent to Owner must be greater than $0")
                return
            
            # Update calculations one final time
            self.update_calculations()
            
            # Show success message
            messagebox.showinfo("Calculation Complete", 
                              f"Rent calculation completed for {self.hoh_var.get()}.\n\n"
                              f"HAP to Owner: ${self.current_calculation['hap_to_owner']:,.0f}\n"
                              f"Tenant Rent: ${self.current_calculation['tenant_rent']:,.0f}\n\n"
                              f"Use 'Export to Excel' to save the results.")
            
            self.status_var.set("✓ Calculation completed successfully")
            
        except ValueError:
            messagebox.showerror("Validation Error", 
                               "Please enter valid numeric values for all financial fields")
    
    def export_to_excel(self):
        """Export calculation to Excel with professional formatting"""
        if not self.current_calculation:
            messagebox.showerror("Export Error", 
                               "Please complete a calculation before exporting")
            return
        
        # Get save location
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save PSH Rent Calculation",
            initialname=f"PSH_Calculation_{self.current_calculation['hoh_name'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}"
        )
        
        if not filename:
            return
        
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PSH Rent Calculation"
            
            # Define styles
            header_font = Font(size=16, bold=True, color='2E7D32')
            section_font = Font(size=12, bold=True, color='1976D2')
            label_font = Font(size=10, bold=True)
            value_font = Font(size=10)
            warning_font = Font(size=10, bold=True, color='D32F2F')
            success_font = Font(size=10, bold=True, color='388E3C')
            
            header_fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
            section_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
            warning_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Header
            row = 1
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws[f'A{row}']
            cell.value = "PSH RENT CALCULATION REPORT"
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            
            row += 2
            
            # Calculation date and staff
            ws[f'A{row}'] = "Calculation Date:"
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = self.current_calculation['calculation_date']
            
            row += 2
            
            # Household Information
            ws[f'A{row}'] = "HOUSEHOLD INFORMATION"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            row += 1
            
            household_data = [
                ("Head of Household:", self.current_calculation['hoh_name']),
                ("Voucher Bedroom Size:", self.current_calculation['voucher_size']),
                ("Bedrooms in Unit:", self.current_calculation['bedrooms']),
                ("FMR Bedroom Size Used:", self.current_calculation['fmr_bedrooms'])
            ]
            
            for label, value in household_data:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = label_font
                ws[f'B{row}'] = value
                row += 1
            
            row += 1
            
            # Financial Information
            ws[f'A{row}'] = "FINANCIAL INFORMATION"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            row += 1
            
            financial_data = [
                ("Rent to Owner:", f"${self.current_calculation['rent_to_owner']:,.0f}"),
                ("Utility Allowance:", f"${self.current_calculation['utility_allowance']:,.0f}"),
                ("Gross Rent:", f"${self.current_calculation['gross_rent']:,.0f}"),
                ("Total Tenant Payment (TTP):", f"${self.current_calculation['ttp']:,.0f}"),
                ("Fair Market Rent:", f"${self.current_calculation['fmr']:,.0f}")
            ]
            
            for label, value in financial_data:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = label_font
                ws[f'B{row}'] = value
                row += 1
            
            # FMR Warning
            if self.current_calculation['over_fmr'] > 0:
                ws[f'A{row}'] = "⚠️ OVER FMR:"
                ws[f'A{row}'].font = warning_font
                ws[f'B{row}'] = f"${self.current_calculation['over_fmr']:,.0f} - SUPERVISOR APPROVAL REQUIRED"
                ws[f'B{row}'].font = warning_font
                ws[f'A{row}'].fill = warning_fill
                ws[f'B{row}'].fill = warning_fill
                row += 1
            
            row += 1
            
            # Family Composition
            ws[f'A{row}'] = "FAMILY COMPOSITION"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            row += 1
            
            family_data = [
                ("Eligible Members:", self.current_calculation['eligible_members']),
                ("Ineligible Members:", self.current_calculation['ineligible_members']),
                ("Total Family Members:", self.current_calculation['total_members'])
            ]
            
            for label, value in family_data:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = label_font
                ws[f'B{row}'] = value
                row += 1
            
            # Mixed family info
            if self.current_calculation['is_mixed_family']:
                ws[f'A{row}'] = "Mixed Family Proration:"
                ws[f'A{row}'].font = warning_font
                ws[f'B{row}'] = f"{self.current_calculation['prorate_percentage']:.1%}"
                ws[f'B{row}'].font = warning_font
                row += 1
            
            row += 1
            
            # CALCULATION RESULTS (highlighted)
            ws[f'A{row}'] = "CALCULATION RESULTS"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            row += 1
            
            # Key results
            results_data = [
                ("HAP to Owner:", f"${self.current_calculation['hap_to_owner']:,.0f}"),
                ("Tenant Rent:", f"${self.current_calculation['tenant_rent']:,.0f}"),
                ("Utility Reimbursement:", f"${self.current_calculation['utility_reimbursement']:,.0f}")
            ]
            
            for label, value in results_data:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(size=12, bold=True)
                ws[f'B{row}'] = value
                ws[f'B{row}'].font = Font(size=12, bold=True, color='2E7D32')
                row += 1
            
            # Mixed family results
            if self.current_calculation['is_mixed_family']:
                row += 1
                ws[f'A{row}'] = "PRORATED ASSISTANCE (Mixed Family)"
                ws[f'A{row}'].font = section_font
                ws[f'A{row}'].fill = warning_fill
                row += 1
                
                ws[f'A{row}'] = "Prorated HAP to Owner:"
                ws[f'A{row}'].font = label_font
                ws[f'B{row}'] = f"${self.current_calculation['prorated_hap']:,.0f}"
                ws[f'B{row}'].font = warning_font
                row += 1
                
                ws[f'A{row}'] = "Mixed Family Rent:"
                ws[f'A{row}'].font = label_font
                ws[f'B{row}'] = f"${self.current_calculation['mixed_family_rent']:,.0f}"
                ws[f'B{row}'].font = warning_font
                row += 1
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 2
                ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
            
            # Save workbook
            wb.save(filename)
            
            messagebox.showinfo("Export Successful", 
                              f"Calculation exported to:\n{filename}\n\n"
                              f"The file is ready for printing or sharing.")
            
            self.status_var.set(f"✓ Exported to {os.path.basename(filename)}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export file:\n{str(e)}")
    
    def clear_form(self):
        """Clear all form fields"""
        if messagebox.askyesno("Clear Form", 
                             "Are you sure you want to clear all fields?"):
            self.hoh_var.set("")
            self.voucher_var.set("0")
            self.bedrooms_var.set("0")
            self.rent_var.set("0")
            self.utility_var.set("0")
            self.ttp_var.set("50")
            self.eligible_var.set("1")
            self.ineligible_var.set("0")
            
            self.current_calculation = {}
            self.update_fmr_display()
            self.update_calculations()
            
            self.status_var.set("Form cleared - ready for new calculation")
    
    def save_fmr_rates(self):
        """Save FMR rates to settings"""
        try:
            for bedroom_count in range(6):
                self.fmr_data[bedroom_count] = float(self.fmr_vars[bedroom_count].get())
            
            self.save_settings()
            self.update_fmr_display()
            
            messagebox.showinfo("Settings Saved", "FMR rates have been updated successfully!")
            self.status_var.set("✓ FMR rates saved")
            
        except ValueError:
            messagebox.showerror("Invalid Input", "Please enter valid numeric values for all FMR rates")
    
    def reset_fmr_rates(self):
        """Reset FMR rates to defaults"""
        if messagebox.askyesno("Reset FMR Rates", 
                             "Reset all FMR rates to default values?"):
            default_fmr = {
                0: 2485, 1: 2977, 2: 3604, 3: 4604, 4: 4772, 5: 5000
            }
            
            for bedroom_count in range(6):
                self.fmr_vars[bedroom_count].set(str(default_fmr[bedroom_count]))
                self.fmr_data[bedroom_count] = default_fmr[bedroom_count]
            
            self.save_settings()
            self.update_fmr_display()
            
            self.status_var.set("✓ FMR rates reset to defaults")
    
    def load_fmr_from_excel(self):
        """Load FMR rates from Excel file"""
        filename = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
            title="Load FMR Rates from Excel"
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            
            # Look for FMR data (simple format: bedroom count in column A, FMR in column B)
            for row in range(1, 20):  # Check first 20 rows
                try:
                    bedroom_val = ws[f'A{row}'].value
                    fmr_val = ws[f'B{row}'].value
                    
                    if isinstance(bedroom_val, (int, float)) and isinstance(fmr_val, (int, float)):
                        bedroom_count = int(bedroom_val)
                        fmr_amount = float(fmr_val)
                        
                        if 0 <= bedroom_count <= 5 and fmr_amount > 0:
                            self.fmr_data[bedroom_count] = fmr_amount
                            self.fmr_vars[bedroom_count].set(str(fmr_amount))
                            
                except (ValueError, TypeError):
                    continue
            
            self.save_settings()
            self.update_fmr_display()
            
            messagebox.showinfo("Import Successful", 
                               f"FMR rates loaded from {os.path.basename(filename)}")
            
            self.status_var.set(f"✓ FMR rates loaded from {os.path.basename(filename)}")
            
        except Exception as e:
            messagebox.showerror("Import Error", 
                                f"Failed to load FMR rates:\n{str(e)}")
    
    def export_settings(self):
        """Export settings to JSON file"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Export Settings"
        )
        
        if not filename:
            return
        
        try:
            settings = {
                'fmr_data': self.fmr_data,
                'export_date': datetime.now().isoformat(),
                'version': '1.0'
            }
            
            with open(filename, 'w') as f:
                json.dump(settings, f, indent=2)
            
            messagebox.showinfo("Export Successful", 
                               f"Settings exported to {os.path.basename(filename)}")
            
        except Exception as e:
            messagebox.showerror("Export Error", 
                                f"Failed to export settings:\n{str(e)}")
    
    def import_settings(self):
        """Import settings from JSON file"""
        filename = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Import Settings"
        )
        
        if not filename:
            return
        
        try:
            with open(filename, 'r') as f:
                settings = json.load(f)
            
            if 'fmr_data' in settings:
                self.fmr_data = {int(k): float(v) for k, v in settings['fmr_data'].items()}
                
                # Update UI
                for bedroom_count in range(6):
                    if bedroom_count in self.fmr_data:
                        self.fmr_vars[bedroom_count].set(str(self.fmr_data[bedroom_count]))
                
                self.save_settings()
                self.update_fmr_display()
                
                messagebox.showinfo("Import Successful", 
                                   f"Settings imported from {os.path.basename(filename)}")
            else:
                messagebox.showerror("Import Error", "Invalid settings file format")
                
        except Exception as e:
            messagebox.showerror("Import Error", 
                                f"Failed to import settings:\n{str(e)}")
    
    def load_settings(self):
        """Load settings from file"""
        settings_file = os.path.expanduser("~/.psh_calculator_settings.json")
        
        if os.path.exists(settings_file):
            try:
                with open(settings_file, 'r') as f:
                    settings = json.load(f)
                
                if 'fmr_data' in settings:
                    self.fmr_data = {int(k): float(v) for k, v in settings['fmr_data'].items()}
                    
            except Exception:
                pass  # Use defaults if loading fails
    
    def save_settings(self):
        """Save settings to file"""
        settings_file = os.path.expanduser("~/.psh_calculator_settings.json")
        
        try:
            settings = {
                'fmr_data': self.fmr_data,
                'last_updated': datetime.now().isoformat()
            }
            
            with open(settings_file, 'w') as f:
                json.dump(settings, f, indent=2)
                
        except Exception:
            pass  # Fail silently


def main():
    """Main application entry point"""
    root = tk.Tk()
    app = PSHCalculatorApp(root)
    
    # Center window on screen
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f"{width}x{height}+{x}+{y}")
    
    # Start main loop
    root.mainloop()


if __name__ == "__main__":
    main()