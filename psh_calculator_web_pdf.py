#!/usr/bin/env python3
"""
PSH Rent Calculator - Web Version with PDF Generation
Beautiful, responsive web interface with PDF report generation
"""

import json
import os
from datetime import datetime
import webbrowser
import threading
from http.server import HTTPServer, BaseHTTPRequestHandler
import urllib.parse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import tempfile

# PDF generation imports
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.platypus.flowables import HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.graphics.shapes import Drawing, Rect
from reportlab.graphics import renderPDF

class PSHCalculatorWeb:
    def __init__(self):
        self.load_fmr_data()
        self.port = 8080
        
    def load_fmr_data(self):
        """Load FMR data from JSON file"""
        self.config_file = "fmr_config.json"
        self.current_year = datetime.now().year
        
        default_fmr = {
            "year": self.current_year,
            "fmr_rates": {
                "0_bedroom": 1200,
                "1_bedroom": 1400,
                "2_bedroom": 1800,
                "3_bedroom": 2400,
                "4_bedroom": 3000
            },
            "last_updated": datetime.now().isoformat()
        }
        
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r') as f:
                    self.fmr_data = json.load(f)
                    if self.fmr_data.get("year", 0) < self.current_year:
                        self.fmr_data["year"] = self.current_year
                        self.save_fmr_data()
            except:
                self.fmr_data = default_fmr
                self.save_fmr_data()
        else:
            self.fmr_data = default_fmr
            self.save_fmr_data()
    
    def save_fmr_data(self):
        """Save FMR data to JSON file"""
        self.fmr_data["last_updated"] = datetime.now().isoformat()
        with open(self.config_file, 'w') as f:
            json.dump(self.fmr_data, f, indent=4)
    
    def get_html(self):
        """Generate the main HTML interface"""
        return f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>PSH Rent Calculator {self.fmr_data['year']}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }}
        
        .container {{
            max-width: 800px;
            margin: 0 auto;
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.1);
            overflow: hidden;
        }}
        
        .header {{
            background: linear-gradient(135deg, #2c3e50 0%, #34495e 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}
        
        .header h1 {{
            font-size: 2.5em;
            margin-bottom: 10px;
            font-weight: 300;
        }}
        
        .year-badge {{
            display: inline-block;
            background: #3498db;
            padding: 8px 16px;
            border-radius: 20px;
            font-size: 0.9em;
            font-weight: bold;
        }}
        
        .content {{
            padding: 40px;
        }}
        
        .form-group {{
            margin-bottom: 25px;
        }}
        
        label {{
            display: block;
            margin-bottom: 8px;
            font-weight: 600;
            color: #2c3e50;
            font-size: 1.1em;
        }}
        
        input, select {{
            width: 100%;
            padding: 15px;
            border: 2px solid #e0e6ed;
            border-radius: 10px;
            font-size: 1.1em;
            transition: all 0.3s ease;
        }}
        
        input:focus, select:focus {{
            outline: none;
            border-color: #3498db;
            box-shadow: 0 0 20px rgba(52,152,219,0.2);
        }}
        
        .fmr-display {{
            background: linear-gradient(135deg, #27ae60 0%, #2ecc71 100%);
            color: white;
            padding: 20px;
            border-radius: 10px;
            text-align: center;
            font-size: 1.8em;
            font-weight: bold;
            margin: 20px 0;
            box-shadow: 0 10px 30px rgba(39,174,96,0.3);
        }}
        
        .calc-button {{
            background: linear-gradient(135deg, #e74c3c 0%, #c0392b 100%);
            color: white;
            border: none;
            padding: 18px 40px;
            font-size: 1.2em;
            font-weight: bold;
            border-radius: 50px;
            cursor: pointer;
            width: 100%;
            margin: 20px 0;
            transition: all 0.3s ease;
            box-shadow: 0 10px 30px rgba(231,76,60,0.3);
        }}
        
        .calc-button:hover {{
            transform: translateY(-2px);
            box-shadow: 0 15px 40px rgba(231,76,60,0.4);
        }}
        
        .results {{
            background: #f8f9fa;
            border-radius: 10px;
            padding: 30px;
            margin: 20px 0;
            border-left: 5px solid #3498db;
        }}
        
        .result-item {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 15px 0;
            border-bottom: 1px solid #e0e6ed;
            font-size: 1.1em;
        }}
        
        .result-item:last-child {{
            border-bottom: none;
            font-size: 1.3em;
            font-weight: bold;
            color: #e74c3c;
        }}
        
        .export-section {{
            display: flex;
            gap: 10px;
            margin-top: 20px;
            flex-wrap: wrap;
        }}
        
        .export-button {{
            background: #95a5a6;
            color: white;
            border: none;
            padding: 12px 20px;
            border-radius: 25px;
            cursor: pointer;
            font-size: 0.9em;
            transition: all 0.3s ease;
            flex: 1;
            min-width: 120px;
        }}
        
        .export-button:hover {{
            background: #7f8c8d;
            transform: translateY(-1px);
        }}
        
        .export-button.pdf {{
            background: #e74c3c;
        }}
        
        .export-button.pdf:hover {{
            background: #c0392b;
        }}
        
        .admin-button {{
            background: #9b59b6;
            color: white;
            border: none;
            padding: 12px 24px;
            border-radius: 25px;
            cursor: pointer;
            font-size: 1em;
            position: absolute;
            top: 20px;
            right: 20px;
            transition: all 0.3s ease;
        }}
        
        .admin-button:hover {{
            background: #8e44ad;
        }}
        
        .admin-modal {{
            display: none;
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: rgba(0,0,0,0.5);
            z-index: 1000;
        }}
        
        .admin-content {{
            position: absolute;
            top: 50%;
            left: 50%;
            transform: translate(-50%, -50%);
            background: white;
            padding: 40px;
            border-radius: 15px;
            max-width: 500px;
            width: 90%;
            max-height: 80vh;
            overflow-y: auto;
        }}
        
        .close-button {{
            position: absolute;
            top: 15px;
            right: 20px;
            background: none;
            border: none;
            font-size: 1.5em;
            cursor: pointer;
            color: #95a5a6;
        }}
        
        @media (max-width: 768px) {{
            .container {{
                margin: 10px;
                border-radius: 15px;
            }}
            
            .header h1 {{
                font-size: 2em;
            }}
            
            .content {{
                padding: 20px;
            }}
            
            .export-section {{
                flex-direction: column;
            }}
            
            .export-button {{
                flex: none;
                width: 100%;
            }}
            
            .admin-button {{
                position: relative;
                top: auto;
                right: auto;
                width: 100%;
                margin-top: 20px;
            }}
        }}
    </style>
</head>
<body>
    <button class="admin-button" onclick="openAdmin()">⚙️ Admin</button>
    
    <div class="container">
        <div class="header">
            <h1>PSH Rent Calculator</h1>
            <div class="year-badge">Fair Market Rent Data - {self.fmr_data['year']}</div>
        </div>
        
        <div class="content">
            <div class="form-group">
                <label for="address">Property Address</label>
                <input type="text" id="address" placeholder="Enter property address...">
            </div>
            
            <div class="form-group">
                <label for="bedrooms">Number of Bedrooms</label>
                <select id="bedrooms" onchange="updateFMR()">
                    <option value="0">Studio (0 Bedrooms)</option>
                    <option value="1" selected>1 Bedroom</option>
                    <option value="2">2 Bedrooms</option>
                    <option value="3">3 Bedrooms</option>
                    <option value="4">4+ Bedrooms</option>
                </select>
            </div>
            
            <div class="fmr-display" id="fmr-display">
                Fair Market Rent: $1,400.00
            </div>
            
            <div class="form-group">
                <label for="tenant-rent">Tenant Rent Payment</label>
                <input type="number" id="tenant-rent" placeholder="Enter tenant's monthly rent payment..." step="0.01">
            </div>
            
            <button class="calc-button" onclick="calculatePSH()">
                Calculate Project-Based Subsidy
            </button>
            
            <div id="results" style="display: none;" class="results">
                <div class="result-item">
                    <span>Property Address:</span>
                    <span id="result-address"></span>
                </div>
                <div class="result-item">
                    <span>Bedrooms:</span>
                    <span id="result-bedrooms"></span>
                </div>
                <div class="result-item">
                    <span>Fair Market Rent:</span>
                    <span id="result-fmr"></span>
                </div>
                <div class="result-item">
                    <span>Tenant Rent Payment:</span>
                    <span id="result-tenant"></span>
                </div>
                <div class="result-item">
                    <span>PROJECT-BASED SUBSIDY:</span>
                    <span id="result-psh"></span>
                </div>
                
                <div class="export-section">
                    <button class="export-button pdf" onclick="exportPDF()">📄 Print/Save PDF</button>
                    <button class="export-button" onclick="exportExcel()">📊 Export Excel</button>
                    <button class="export-button" onclick="printResults()">🖨️ Quick Print</button>
                </div>
            </div>
        </div>
    </div>
    
    <!-- Admin Modal -->
    <div id="admin-modal" class="admin-modal">
        <div class="admin-content">
            <button class="close-button" onclick="closeAdmin()">&times;</button>
            <h2>FMR Administration</h2>
            <br>
            
            <div class="form-group">
                <label for="admin-year">FMR Year:</label>
                <input type="number" id="admin-year" value="{self.fmr_data['year']}">
            </div>
            
            <div class="form-group">
                <label for="fmr-0">Studio (0 BR) FMR:</label>
                <input type="number" id="fmr-0" value="{self.fmr_data['fmr_rates']['0_bedroom']}" step="0.01">
            </div>
            
            <div class="form-group">
                <label for="fmr-1">1 Bedroom FMR:</label>
                <input type="number" id="fmr-1" value="{self.fmr_data['fmr_rates']['1_bedroom']}" step="0.01">
            </div>
            
            <div class="form-group">
                <label for="fmr-2">2 Bedroom FMR:</label>
                <input type="number" id="fmr-2" value="{self.fmr_data['fmr_rates']['2_bedroom']}" step="0.01">
            </div>
            
            <div class="form-group">
                <label for="fmr-3">3 Bedroom FMR:</label>
                <input type="number" id="fmr-3" value="{self.fmr_data['fmr_rates']['3_bedroom']}" step="0.01">
            </div>
            
            <div class="form-group">
                <label for="fmr-4">4+ Bedroom FMR:</label>
                <input type="number" id="fmr-4" value="{self.fmr_data['fmr_rates']['4_bedroom']}" step="0.01">
            </div>
            
            <button class="calc-button" onclick="saveAdmin()">Save Settings</button>
        </div>
    </div>
    
    <script>
        let fmrRates = {json.dumps(self.fmr_data['fmr_rates'])};
        let currentYear = {self.fmr_data['year']};
        let lastCalculation = null;
        
        function updateFMR() {{
            const bedrooms = document.getElementById('bedrooms').value;
            const key = bedrooms + '_bedroom';
            const fmr = fmrRates[key] || 0;
            document.getElementById('fmr-display').innerHTML = 
                `Fair Market Rent: $` + fmr.toLocaleString('en-US', {{minimumFractionDigits: 2}});
        }}
        
        function calculatePSH() {{
            const address = document.getElementById('address').value.trim();
            const bedrooms = document.getElementById('bedrooms').value;
            const tenantRent = parseFloat(document.getElementById('tenant-rent').value) || 0;
            
            if (!address) {{
                alert('Please enter a property address');
                return;
            }}
            
            const key = bedrooms + '_bedroom';
            const fmr = fmrRates[key] || 0;
            const psh = Math.max(0, fmr - tenantRent);
            
            // Store calculation
            lastCalculation = {{
                address: address,
                bedrooms: bedrooms,
                fmr: fmr,
                tenantRent: tenantRent,
                psh: psh,
                year: currentYear,
                date: new Date().toLocaleString()
            }};
            
            // Display results
            document.getElementById('result-address').textContent = address;
            document.getElementById('result-bedrooms').textContent = 
                bedrooms === '0' ? 'Studio' : bedrooms + (bedrooms === '4' ? '+' : '');
            document.getElementById('result-fmr').textContent = 
                '$' + fmr.toLocaleString('en-US', {{minimumFractionDigits: 2}});
            document.getElementById('result-tenant').textContent = 
                '$' + tenantRent.toLocaleString('en-US', {{minimumFractionDigits: 2}});
            document.getElementById('result-psh').textContent = 
                '$' + psh.toLocaleString('en-US', {{minimumFractionDigits: 2}});
            
            document.getElementById('results').style.display = 'block';
            document.getElementById('results').scrollIntoView({{ behavior: 'smooth' }});
        }}
        
        function exportPDF() {{
            if (!lastCalculation) {{
                alert('Please calculate PSH first');
                return;
            }}
            
            fetch('/pdf', {{
                method: 'POST',
                headers: {{ 'Content-Type': 'application/json' }},
                body: JSON.stringify(lastCalculation)
            }})
            .then(response => response.blob())
            .then(blob => {{
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.style.display = 'none';
                a.href = url;
                a.download = 'PSH_Calculation_' + lastCalculation.address.replace(/[^a-zA-Z0-9]/g, '_') + '.pdf';
                document.body.appendChild(a);
                a.click();
                window.URL.revokeObjectURL(url);
                alert('PDF report downloaded successfully!');
            }})
            .catch(error => {{
                console.error('PDF export error:', error);
                alert('PDF export failed. Please try again.');
            }});
        }}
        
        function exportExcel() {{
            if (!lastCalculation) {{
                alert('Please calculate PSH first');
                return;
            }}
            
            fetch('/export', {{
                method: 'POST',
                headers: {{ 'Content-Type': 'application/json' }},
                body: JSON.stringify(lastCalculation)
            }})
            .then(response => response.blob())
            .then(blob => {{
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.style.display = 'none';
                a.href = url;
                a.download = 'PSH_Calculation_' + new Date().toISOString().split('T')[0] + '.xlsx';
                document.body.appendChild(a);
                a.click();
                window.URL.revokeObjectURL(url);
                alert('Excel file downloaded successfully!');
            }})
            .catch(error => {{
                console.error('Export error:', error);
                alert('Export failed. Please try again.');
            }});
        }}
        
        function printResults() {{
            if (!lastCalculation) {{
                alert('Please calculate PSH first');
                return;
            }}
            
            const printWindow = window.open('', '_blank');
            printWindow.document.write(`
                <html>
                <head>
                    <title>PSH Calculation Report</title>
                    <style>
                        body {{ font-family: Arial, sans-serif; margin: 20px; }}
                        h1 {{ color: #2c3e50; text-align: center; }}
                        .result {{ margin: 10px 0; padding: 10px; border-bottom: 1px solid #ccc; }}
                        .final {{ font-weight: bold; font-size: 1.2em; color: #e74c3c; }}
                    </style>
                </head>
                <body>
                    <h1>PSH Calculation Report</h1>
                    <p><strong>Generated:</strong> ${{lastCalculation.date}}</p>
                    <p><strong>FMR Year:</strong> ${{lastCalculation.year}}</p>
                    <br>
                    <div class="result"><strong>Property Address:</strong> ${{lastCalculation.address}}</div>
                    <div class="result"><strong>Bedrooms:</strong> ${{lastCalculation.bedrooms === '0' ? 'Studio' : lastCalculation.bedrooms + (lastCalculation.bedrooms === '4' ? '+' : '')}}</div>
                    <div class="result"><strong>Fair Market Rent:</strong> $${{lastCalculation.fmr.toLocaleString('en-US', {{minimumFractionDigits: 2}})}}</div>
                    <div class="result"><strong>Tenant Rent Payment:</strong> $${{lastCalculation.tenantRent.toLocaleString('en-US', {{minimumFractionDigits: 2}})}}</div>
                    <div class="result final"><strong>PROJECT-BASED SUBSIDY:</strong> $${{lastCalculation.psh.toLocaleString('en-US', {{minimumFractionDigits: 2}})}}</div>
                </body>
                </html>
            `);
            printWindow.document.close();
            printWindow.print();
        }}
        
        function openAdmin() {{
            document.getElementById('admin-modal').style.display = 'block';
        }}
        
        function closeAdmin() {{
            document.getElementById('admin-modal').style.display = 'none';
        }}
        
        function saveAdmin() {{
            const adminData = {{
                year: parseInt(document.getElementById('admin-year').value),
                fmr_rates: {{
                    '0_bedroom': parseFloat(document.getElementById('fmr-0').value),
                    '1_bedroom': parseFloat(document.getElementById('fmr-1').value),
                    '2_bedroom': parseFloat(document.getElementById('fmr-2').value),
                    '3_bedroom': parseFloat(document.getElementById('fmr-3').value),
                    '4_bedroom': parseFloat(document.getElementById('fmr-4').value)
                }}
            }};
            
            fetch('/admin', {{
                method: 'POST',
                headers: {{ 'Content-Type': 'application/json' }},
                body: JSON.stringify(adminData)
            }})
            .then(response => response.json())
            .then(data => {{
                if (data.success) {{
                    fmrRates = adminData.fmr_rates;
                    currentYear = adminData.year;
                    updateFMR();
                    document.querySelector('.year-badge').textContent = 
                        'Fair Market Rent Data - ' + adminData.year;
                    closeAdmin();
                    alert('Settings saved successfully!');
                    setTimeout(() => location.reload(), 500);
                }} else {{
                    alert('Error saving settings: ' + data.error);
                }}
            }})
            .catch(error => {{
                console.error('Admin save error:', error);
                alert('Save failed. Please try again.');
            }});
        }}
        
        // Initialize FMR display
        updateFMR();
        
        // Click outside modal to close
        document.getElementById('admin-modal').addEventListener('click', function(e) {{
            if (e.target === this) {{
                closeAdmin();
            }}
        }});
    </script>
</body>
</html>'''

class PSHRequestHandler(BaseHTTPRequestHandler):
    def __init__(self, calculator, *args, **kwargs):
        self.calculator = calculator
        super().__init__(*args, **kwargs)
    
    def do_GET(self):
        if self.path == '/' or self.path == '/index.html':
            self.send_response(200)
            self.send_header('Content-type', 'text/html')
            self.end_headers()
            self.wfile.write(self.calculator.get_html().encode())
        else:
            self.send_response(404)
            self.end_headers()
    
    def do_POST(self):
        content_length = int(self.headers['Content-Length'])
        post_data = self.rfile.read(content_length)
        
        if self.path == '/admin':
            try:
                data = json.loads(post_data.decode())
                self.calculator.fmr_data['year'] = data['year']
                self.calculator.fmr_data['fmr_rates'] = data['fmr_rates']
                self.calculator.save_fmr_data()
                
                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"success": True}).encode())
                
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"success": False, "error": str(e)}).encode())
        
        elif self.path == '/pdf':
            try:
                data = json.loads(post_data.decode())
                pdf_file = self.create_pdf_report(data)
                
                self.send_response(200)
                self.send_header('Content-type', 'application/pdf')
                self.send_header('Content-Disposition', 'attachment; filename="PSH_Calculation.pdf"')
                self.end_headers()
                
                with open(pdf_file, 'rb') as f:
                    self.wfile.write(f.read())
                
                os.unlink(pdf_file)  # Clean up temp file
                
            except Exception as e:
                self.send_response(500)
                self.end_headers()
        
        elif self.path == '/export':
            try:
                data = json.loads(post_data.decode())
                excel_file = self.create_excel_report(data)
                
                self.send_response(200)
                self.send_header('Content-type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', 'attachment; filename="PSH_Calculation.xlsx"')
                self.end_headers()
                
                with open(excel_file, 'rb') as f:
                    self.wfile.write(f.read())
                
                os.unlink(excel_file)  # Clean up temp file
                
            except Exception as e:
                self.send_response(500)
                self.end_headers()
        else:
            self.send_response(404)
            self.end_headers()
    
    def create_pdf_report(self, calculation_data):
        """Create professional PDF report"""
        temp_file = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
        
        # Create PDF document
        doc = SimpleDocTemplate(
            temp_file.name,
            pagesize=letter,
            topMargin=1*inch,
            bottomMargin=1*inch,
            leftMargin=1*inch,
            rightMargin=1*inch
        )
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#34495e'),
            spaceBefore=20,
            spaceAfter=10
        )
        
        # Build content
        content = []
        
        # Title
        title = Paragraph("PSH RENT CALCULATION REPORT", title_style)
        content.append(title)
        
        # Year and date
        year_info = Paragraph(
            f"Fair Market Rent Year: {calculation_data['year']} | Generated: {calculation_data['date']}",
            styles['Normal']
        )
        content.append(year_info)
        content.append(Spacer(1, 20))
        
        # Horizontal line
        content.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#3498db')))
        content.append(Spacer(1, 20))
        
        # Property details table
        bedroom_display = calculation_data['bedrooms']
        if bedroom_display == '0':
            bedroom_display = 'Studio'
        elif bedroom_display == '4':
            bedroom_display = '4+'
        
        data = [
            ['Property Details', 'Value'],
            ['Property Address', calculation_data['address']],
            ['Number of Bedrooms', bedroom_display],
            ['Fair Market Rent (FMR)', f"${calculation_data['fmr']:,.2f}"],
            ['Tenant Rent Payment', f"${calculation_data['tenantRent']:,.2f}"],
            ['', ''],
            ['PROJECT-BASED SUBSIDY (PSH)', f"${calculation_data['psh']:,.2f}"]
        ]
        
        table = Table(data, colWidths=[3*inch, 2.5*inch])
        table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 11),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            
            # PSH row (last row)
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e74c3c')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 14),
            ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
            
            # Spacing row (second to last)
            ('BACKGROUND', (0, -2), (-1, -2), colors.white),
            ('GRID', (0, -2), (-1, -2), 0, colors.white),
        ]))
        
        content.append(table)
        content.append(Spacer(1, 30))
        
        # Calculation explanation
        explanation_style = ParagraphStyle(
            'Explanation',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#7f8c8d'),
            leftIndent=20,
            rightIndent=20
        )
        
        explanation = Paragraph(
            f"<b>Calculation Method:</b> Project-Based Subsidy (PSH) = Fair Market Rent - Tenant Rent Payment<br/>"
            f"<b>FMR Source:</b> {calculation_data['year']} Fair Market Rent data<br/>"
            f"<b>Property Type:</b> {bedroom_display} unit<br/>"
            f"<b>Subsidy Amount:</b> The property owner will receive ${calculation_data['psh']:,.2f} monthly from the housing assistance program.",
            explanation_style
        )
        content.append(explanation)
        
        # Footer
        content.append(Spacer(1, 40))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#95a5a6'),
            alignment=1  # Center alignment
        )
        
        footer = Paragraph(
            "This calculation is based on current Fair Market Rent data and provided tenant rent amount.<br/>"
            "For questions or support, contact: monday@aimagery.com",
            footer_style
        )
        content.append(footer)
        
        # Build PDF
        doc.build(content)
        
        return temp_file.name
    
    def create_excel_report(self, calculation_data):
        """Create Excel report and return filename"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PSH Calculation"
        
        # Styles
        header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        title_font = Font(name='Calibri', size=18, bold=True, color='2F5597')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Title
        ws.merge_cells('A1:C1')
        ws['A1'] = 'PSH RENT CALCULATION REPORT'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Date and year
        ws.merge_cells('A2:C2')
        ws['A2'] = f"FMR Year: {calculation_data['year']} | Generated: {calculation_data['date']}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Data
        row = 4
        data_rows = [
            ('Property Address', calculation_data['address']),
            ('Bedrooms', calculation_data['bedrooms']),
            ('Fair Market Rent (FMR)', f"${calculation_data['fmr']:,.2f}"),
            ('Tenant Rent Payment', f"${calculation_data['tenantRent']:,.2f}"),
            ('PROJECT-BASED SUBSIDY (PSH)', f"${calculation_data['psh']:,.2f}")
        ]
        
        for i, (label, value) in enumerate(data_rows, row):
            ws.cell(row=i, column=1, value=label).border = border
            cell = ws.cell(row=i, column=2, value=value)
            cell.border = border
            if 'PSH' in label:
                cell.font = Font(bold=True, color='2F5597')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        
        # Save to temp file
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(temp_file.name)
        return temp_file.name
    
    def log_message(self, format, *args):
        pass  # Suppress log messages

def make_handler(calculator):
    def handler(*args, **kwargs):
        return PSHRequestHandler(calculator, *args, **kwargs)
    return handler

def main():
    calculator = PSHCalculatorWeb()
    
    # Find available port
    port = 8080
    while port < 8090:
        try:
            server = HTTPServer(('localhost', port), make_handler(calculator))
            calculator.port = port
            break
        except OSError:
            port += 1
    else:
        print("Could not find available port")
        return
    
    print(f"╔════════════════════════════════════════╗")
    print(f"║       PSH RENT CALCULATOR - 2026      ║")
    print(f"║        Web Edition with PDF            ║")
    print(f"╚════════════════════════════════════════╝")
    print(f"")
    print(f"🌐 Starting server on http://localhost:{port}")
    print(f"📊 FMR Year: {calculator.fmr_data['year']}")
    print(f"📄 PDF Reports: Professional printable forms")
    print(f"⚙️  Admin controls available in interface")
    print(f"")
    print(f"Opening browser...")
    
    # Open browser
    def open_browser():
        webbrowser.open(f'http://localhost:{port}')
    
    threading.Timer(1.0, open_browser).start()
    
    try:
        print(f"✅ Server running! Close this window to stop the calculator.")
        server.serve_forever()
    except KeyboardInterrupt:
        print(f"\n👋 Shutting down server...")
        server.shutdown()
        
if __name__ == "__main__":
    main()